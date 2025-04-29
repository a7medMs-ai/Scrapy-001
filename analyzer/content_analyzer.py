import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

def generate_excel_report(data_dict, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for language, records in data_dict.items():
            df = pd.DataFrame(records)

            # Reorder columns for clarity
            column_order = [
                "Page Counter",
                "Page Name",
                "Word Count",
                "Segments",
                "Has Media",
                "Content Type",
                "URL",
                "Page Content"
            ]
            df = df[column_order]

            # Add total row
            total_row = {
                "Page Counter": "Total",
                "Word Count": df["Word Count"].sum(),
                "Segments": df["Segments"].sum(),
                "Page Name": "",
                "Has Media": "",
                "Content Type": "",
                "URL": "",
                "Page Content": ""
            }
            df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

            # Write to sheet
            df.to_excel(writer, sheet_name=language.upper(), index=False)

        writer.save()

    # Styling (bold last row)
    wb = load_workbook(output_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_row = ws.max_row
        for cell in ws[max_row]:
            cell.font = Font(bold=True)
    wb.save(output_path)
